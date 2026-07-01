import pandas as pd
from datetime import datetime
import openpyxl
import shutil
import zipfile
import re
import os
from openpyxl.styles import Alignment, Border, Side, Font
import customtkinter as ctk
from tkinter import filedialog, messagebox

# ==========================================
# 1. CORE LOGIC (Dari app.py lama)
# ==========================================

def strip_data_validations(excel_path):
    temp_path = excel_path + ".tmp"
    with zipfile.ZipFile(excel_path, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w') as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/'):
                    content_str = content.decode('utf-8')
                    content_str = re.sub(r'<dataValidations.*?</dataValidations>', '', content_str, flags=re.DOTALL)
                    content = content_str.encode('utf-8')
                zout.writestr(item, content)
    shutil.move(temp_path, excel_path)

def generate_laporan(file_csv, template_excel, file_output, nama_pegawai, opd, projek, role):
    df = pd.read_csv(file_csv, sep=';')
    data_saya = df[df['nama'].str.lower() == nama_pegawai.lower()]

    hasil_rekap = []
    if data_saya.empty:
        return False, f"Maaf, nama '{nama_pegawai}' tidak ditemukan di file CSV."

    kolom_tanggal = df.columns[2:]

    for tanggal_str in kolom_tanggal:
        absen = data_saya.iloc[0][tanggal_str]
        tanggal_str_bersih = str(tanggal_str).strip()
        
        try:
            # Standarisasi format tanggal (ubah / menjadi -)
            tanggal_str_bersih = tanggal_str_bersih.replace('/', '-')
            tanggal_obj = datetime.strptime(tanggal_str_bersih, '%d-%m-%Y')
            nama_hari = tanggal_obj.strftime('%A')
            tanggal_format_baru = tanggal_obj.strftime('%Y-%m-%d')
        except ValueError:
            continue

        jam_masuk = '-'
        jam_pulang = '-'
        durasi_kerja = '-'
        keterangan = '-'

        if nama_hari == 'Saturday':
            keterangan = 'Sabtu'
        elif nama_hari == 'Sunday':
            keterangan = 'Minggu'
        else:
            if pd.notna(absen) and str(absen).strip() != '':
                try:
                    if ' - ' in str(absen):
                        jam_masuk, jam_pulang = str(absen).split(' - ')
                        jam_masuk = jam_masuk.replace('.', ':').strip()
                        jam_pulang = jam_pulang.replace('.', ':').strip()
                        waktu_masuk = datetime.strptime(jam_masuk, '%H:%M:%S')
                        waktu_pulang = datetime.strptime(jam_pulang, '%H:%M:%S')
                        selisih = waktu_pulang - waktu_masuk
                        jam = selisih.seconds // 3600
                        menit = (selisih.seconds // 60) % 60
                        durasi_kerja = f'{jam} jam {menit} menit'
                        keterangan = 'Hadir'
                    else:
                        jam_masuk = str(absen).strip()
                        keterangan = 'Format Error / Pulang Kosong'
                except Exception as e:
                    keterangan = 'Format Error/Tidak Lengkap'
            else:
                keterangan = 'Tidak Hadir / Cuti'

        hasil_rekap.append({
            'Tanggal(s)': tanggal_format_baru,
            'Jam Masuk': jam_masuk,
            'Jam Pulang': jam_pulang,
            'Durasi Kerja': durasi_kerja,
            'Keterangan': keterangan
        })

    ttd_date = datetime.now()

    shutil.copy(template_excel, file_output)
    strip_data_validations(file_output)

    wb = openpyxl.load_workbook(file_output)
    ws = wb.active

    # Update Biodata
    ws.cell(row=7, column=3, value=opd)
    ws.cell(row=7, column=6, value=nama_pegawai)
    ws.cell(row=8, column=3, value=projek)
    ws.cell(row=8, column=6, value=role)
    
    # Update Bulan/Tahun di Pojok Kanan Atas (Row 5, Col 7)
    if len(kolom_tanggal) > 0:
        first_date_str = str(kolom_tanggal[0]).strip().replace('/', '-')
        try:
            first_date_obj = datetime.strptime(first_date_str, '%d-%m-%Y')
            month_year_str = first_date_obj.strftime("%m/%y")
            ws.cell(row=5, column=7).value = month_year_str
        except ValueError:
            pass

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    regular_font = Font(name='Arial', size=11, bold=False)
    red_font = Font(name='Arial', size=11, bold=False, color="FF0000")
    left_align = Alignment(horizontal='left')

    for r in range(11, 42):
        for c in range(2, 7):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).border = Border()
            ws.cell(row=r, column=c).font = regular_font
            ws.cell(row=r, column=c).alignment = left_align
            
    baris_sekarang = 11
    for data_hari in hasil_rekap:
        ws.cell(row=baris_sekarang, column=2, value=data_hari['Tanggal(s)']).border = thin_border
        ws.cell(row=baris_sekarang, column=3, value=data_hari['Jam Masuk']).border = thin_border
        ws.cell(row=baris_sekarang, column=4, value=data_hari['Jam Pulang']).border = thin_border
        ws.cell(row=baris_sekarang, column=5, value=data_hari['Durasi Kerja']).border = thin_border
        ws.cell(row=baris_sekarang, column=6, value=data_hari['Keterangan']).border = thin_border
        
        keterangan = data_hari['Keterangan']
        is_red = (keterangan in ['Sabtu', 'Minggu', 'Tidak Hadir / Cuti'])
        current_font = red_font if is_red else regular_font

        for c in range(2, 7):
            ws.cell(row=baris_sekarang, column=c).font = current_font
            ws.cell(row=baris_sekarang, column=c).alignment = left_align
            
        baris_sekarang += 1

    last_data_row = baris_sekarang - 1
    
    nama_row = None
    for r in range(last_data_row, 60):
        val = ws.cell(row=r, column=2).value
        if val and 'Nama' in str(val):
            nama_row = r
            break
            
    if nama_row:
        signature_start = nama_row - 1
        target_signature_start = last_data_row + 3
        
        selisih = target_signature_start - signature_start
        if selisih != 0:
            merged_offsets = []
            merged_cells_to_remove = []
            for mc in ws.merged_cells.ranges:
                if mc.min_row >= signature_start and mc.max_row <= signature_start + 15:
                    merged_offsets.append((mc.min_row - signature_start, mc.min_col, mc.max_row - signature_start, mc.max_col))
                    merged_cells_to_remove.append(mc)
            for mc in merged_cells_to_remove:
                ws.merged_cells.remove(mc)

            ws.move_range(f"A{signature_start}:H{signature_start + 10}", rows=selisih, cols=0)

            for r_min, c_min, r_max, c_max in merged_offsets:
                ws.merge_cells(start_row=target_signature_start + r_min, start_column=c_min, end_row=target_signature_start + r_max, end_column=c_max)

        for r in range(last_data_row + 1, target_signature_start):
            for c in range(1, 8):
                ws.cell(row=r, column=c).value = None
                ws.cell(row=r, column=c).border = Border()
                ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(fill_type=None)
                
        # Fix row heights manually for the signature block
        ws.row_dimensions[target_signature_start].height = 20.25      # Tenaga Ahli
        ws.row_dimensions[target_signature_start + 1].height = 21.75  # Nama
        ws.row_dimensions[target_signature_start + 2].height = 47.25  # Ttd
        ws.row_dimensions[target_signature_start + 3].height = 20.25  # Space
        ws.row_dimensions[target_signature_start + 4].height = 20.25  # Space
        ws.row_dimensions[target_signature_start + 5].height = 20.25  # Tanggal
        
        if selisih != 0:
            for r in range(target_signature_start + 6, 60):
                ws.row_dimensions[r].height = 20.25
                
        # Update Tanggal Tanda Tangan & Nama
        for r in range(target_signature_start, target_signature_start + 10):
            val = ws.cell(row=r, column=2).value
            if val and 'Tanggal' in str(val):
                tanggal_str = ttd_date.strftime("%d/%m/%Y")
                ws.cell(row=r, column=3).value = tanggal_str
                ws.cell(row=r, column=6).value = tanggal_str
            if val and 'Nama' in str(val):
                ws.cell(row=r, column=3).value = nama_pegawai

    wb.save(file_output)
    return True, f"Sukses! File disimpan sebagai:\n{os.path.basename(file_output)}"


# ==========================================
# 2. GUI APPLICATION (CustomTkinter)
# ==========================================

ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Generator Laporan Absensi")
        self.geometry("600x550")
        
        # Tetap menggunakan template di folder yang sama (sesuai spesifikasi kita sebelumnya)
        self.csv_path = None
        self.template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Laporan Absensi Maret 2026.xlsx')
        
        # UI Elements
        self.grid_columnconfigure(1, weight=1)

        # 1. Judul
        self.title_label = ctk.CTkLabel(self, text="Form Laporan Absensi", font=ctk.CTkFont(size=20, weight="bold"))
        self.title_label.grid(row=0, column=0, columnspan=3, padx=20, pady=(20, 10))

        # 2. Pilih CSV
        self.lbl_csv = ctk.CTkLabel(self, text="File Data CSV:")
        self.lbl_csv.grid(row=1, column=0, padx=20, pady=10, sticky="e")
        self.entry_csv = ctk.CTkEntry(self, state="disabled")
        self.entry_csv.grid(row=1, column=1, padx=(0, 10), pady=10, sticky="ew")
        self.btn_csv = ctk.CTkButton(self, text="Browse", command=self.browse_csv)
        self.btn_csv.grid(row=1, column=2, padx=20, pady=10)

        # 3. Dropdown Nama Pegawai
        self.lbl_nama = ctk.CTkLabel(self, text="Nama Pegawai:")
        self.lbl_nama.grid(row=2, column=0, padx=20, pady=10, sticky="e")
        self.cmb_nama = ctk.CTkComboBox(self, values=["Pilih CSV Terlebih Dahulu..."])
        self.cmb_nama.grid(row=2, column=1, columnspan=2, padx=20, pady=10, sticky="ew")
        self.cmb_nama.set("Pilih CSV Terlebih Dahulu...")

        # 4. OPD (Default)
        self.lbl_opd = ctk.CTkLabel(self, text="OPD:")
        self.lbl_opd.grid(row=3, column=0, padx=20, pady=10, sticky="e")
        self.entry_opd = ctk.CTkEntry(self)
        self.entry_opd.insert(0, "Dinas Komunikasi dan Informatika")
        self.entry_opd.grid(row=3, column=1, columnspan=2, padx=20, pady=10, sticky="ew")

        # 5. Nama Projek
        self.lbl_projek = ctk.CTkLabel(self, text="Nama Projek:")
        self.lbl_projek.grid(row=4, column=0, padx=20, pady=10, sticky="e")
        self.entry_projek = ctk.CTkEntry(self)
        self.entry_projek.insert(0, "Open SID (Sistem Informasi Desa) Kab. Badung")
        self.entry_projek.grid(row=4, column=1, columnspan=2, padx=20, pady=10, sticky="ew")

        # 6. Role Pegawai
        self.lbl_role = ctk.CTkLabel(self, text="Role Pegawai:")
        self.lbl_role.grid(row=5, column=0, padx=20, pady=10, sticky="e")
        self.entry_role = ctk.CTkEntry(self)
        self.entry_role.insert(0, "Full Stack Web Developer")
        self.entry_role.grid(row=5, column=1, columnspan=2, padx=20, pady=10, sticky="ew")

        # 7. Generate Button
        self.btn_generate = ctk.CTkButton(self, text="Generate Laporan Excel", command=self.generate, font=ctk.CTkFont(size=15, weight="bold"), height=50)
        self.btn_generate.grid(row=6, column=0, columnspan=3, padx=20, pady=30, sticky="ew")

    def browse_csv(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if file_path:
            self.csv_path = file_path
            self.entry_csv.configure(state="normal")
            self.entry_csv.delete(0, 'end')
            self.entry_csv.insert(0, os.path.basename(file_path))
            self.entry_csv.configure(state="disabled")
            
            # Load names dynamically
            try:
                df = pd.read_csv(file_path, sep=';')
                if 'nama' in df.columns:
                    names = df['nama'].dropna().unique().tolist()
                    self.cmb_nama.configure(values=names)
                    if names:
                        self.cmb_nama.set(names[0])
                else:
                    messagebox.showerror("Error", "Kolom 'nama' tidak ditemukan di CSV ini.")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal membaca CSV: {str(e)}")

    def generate(self):
        if not self.csv_path:
            messagebox.showwarning("Peringatan", "Harap pilih file CSV terlebih dahulu!")
            return
            
        nama = self.cmb_nama.get()
        if not nama or nama == "Pilih CSV Terlebih Dahulu...":
            messagebox.showwarning("Peringatan", "Harap pilih nama pegawai yang valid!")
            return
            
        opd = self.entry_opd.get()
        projek = self.entry_projek.get()
        role = self.entry_role.get()
        
        # Extract short name for filename
        words = nama.split(' ')
        if len(words) > 1 and len(words[0]) <= 2:
            first_name = words[1] # Cth: 'I Made' -> 'Made'
        else:
            first_name = words[0]
            
        clean_name = re.sub(r'[^a-zA-Z0-9]', '', first_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"Laporan_Absensi_{clean_name}_{timestamp}.xlsx"
        
        output_path = os.path.join(os.path.dirname(self.csv_path), output_name)

        try:
            success, msg = generate_laporan(
                file_csv=self.csv_path,
                template_excel=self.template_path,
                file_output=output_path,
                nama_pegawai=nama,
                opd=opd,
                projek=projek,
                role=role
            )
            
            if success:
                messagebox.showinfo("Berhasil", msg)
            else:
                messagebox.showerror("Gagal", msg)
        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan saat memproses laporan:\n{str(e)}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
