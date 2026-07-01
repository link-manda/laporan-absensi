import pandas as pd
from datetime import datetime
import openpyxl
import shutil
import zipfile
import re
import os
from openpyxl.styles import Alignment, Border, Side, Font

def strip_data_validations(excel_path):
    temp_path = excel_path + ".tmp"
    with zipfile.ZipFile(excel_path, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w') as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/'):
                    content_str = content.decode('utf-8')
                    content_str = re.sub(r'<dataValidations.*?</dataValidations>', '', content_str, flags=re.DOTALL)
                    content = content_str.encode('utf-8')
                zout.writestr(item, content)
    shutil.move(temp_path, excel_path)

def generate_laporan(df, template_excel, file_output, nama_pegawai, opd, projek, role):
    data_saya = df[df['nama'].str.lower() == nama_pegawai.lower()]

    hasil_rekap = []
    if data_saya.empty:
        return False, f"Maaf, nama '{nama_pegawai}' tidak ditemukan di file CSV."

    kolom_tanggal = df.columns[2:]

    for tanggal_str in kolom_tanggal:
        absen = data_saya.iloc[0][tanggal_str]
        tanggal_str_bersih = str(tanggal_str).strip()
        
        try:
            # Standarisasi format tanggal (ubah / menjadi -)
            tanggal_str_bersih = tanggal_str_bersih.replace('/', '-')
            tanggal_obj = datetime.strptime(tanggal_str_bersih, '%d-%m-%Y')
            nama_hari = tanggal_obj.strftime('%A')
            tanggal_format_baru = tanggal_obj.strftime('%Y-%m-%d')
        except ValueError:
            continue

        jam_masuk = '-'
        jam_pulang = '-'
        durasi_kerja = '-'
        keterangan = '-'

        if nama_hari == 'Saturday':
            keterangan = 'Sabtu'
        elif nama_hari == 'Sunday':
            keterangan = 'Minggu'
        else:
            if pd.notna(absen) and str(absen).strip() != '':
                try:
                    if ' - ' in str(absen):
                        jam_masuk, jam_pulang = str(absen).split(' - ')
                        jam_masuk = jam_masuk.replace('.', ':').strip()
                        jam_pulang = jam_pulang.replace('.', ':').strip()
                        waktu_masuk = datetime.strptime(jam_masuk, '%H:%M:%S')
                        waktu_pulang = datetime.strptime(jam_pulang, '%H:%M:%S')
                        selisih = waktu_pulang - waktu_masuk
                        jam = selisih.seconds // 3600
                        menit = (selisih.seconds // 60) % 60
                        durasi_kerja = f'{jam} jam {menit} menit'
                        keterangan = 'Hadir'
                    else:
                        jam_masuk = str(absen).strip()
                        keterangan = 'Format Error / Pulang Kosong'
                except Exception as e:
                    keterangan = 'Format Error/Tidak Lengkap'
            else:
                keterangan = 'Tidak Hadir / Cuti'

        hasil_rekap.append({
            'Tanggal(s)': tanggal_format_baru,
            'Jam Masuk': jam_masuk,
            'Jam Pulang': jam_pulang,
            'Durasi Kerja': durasi_kerja,
            'Keterangan': keterangan
        })

    ttd_date = datetime.now()

    shutil.copy(template_excel, file_output)
    strip_data_validations(file_output)

    wb = openpyxl.load_workbook(file_output)
    ws = wb.active

    # Update Biodata
    ws.cell(row=7, column=3, value=opd)
    ws.cell(row=7, column=6, value=nama_pegawai)
    ws.cell(row=8, column=3, value=projek)
    ws.cell(row=8, column=6, value=role)

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    regular_font = Font(name='Arial', size=11, bold=False)
    red_font = Font(name='Arial', size=11, bold=False, color="FF0000")
    left_align = Alignment(horizontal='left')

    for r in range(11, 45):
        for c in range(2, 7):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).border = Border()
            ws.cell(row=r, column=c).font = regular_font
            ws.cell(row=r, column=c).alignment = left_align
            
    baris_sekarang = 11
    for data_hari in hasil_rekap:
        ws.cell(row=baris_sekarang, column=2, value=data_hari['Tanggal(s)']).border = thin_border
        ws.cell(row=baris_sekarang, column=3, value=data_hari['Jam Masuk']).border = thin_border
        ws.cell(row=baris_sekarang, column=4, value=data_hari['Jam Pulang']).border = thin_border
        ws.cell(row=baris_sekarang, column=5, value=data_hari['Durasi Kerja']).border = thin_border
        ws.cell(row=baris_sekarang, column=6, value=data_hari['Keterangan']).border = thin_border
        
        keterangan = data_hari['Keterangan']
        is_red = (keterangan in ['Sabtu', 'Minggu', 'Tidak Hadir / Cuti'])
        current_font = red_font if is_red else regular_font

        for c in range(2, 7):
            ws.cell(row=baris_sekarang, column=c).font = current_font
            ws.cell(row=baris_sekarang, column=c).alignment = left_align
            
        baris_sekarang += 1

    last_data_row = baris_sekarang - 1
    
    nama_row = None
    for r in range(last_data_row, last_data_row + 20):
        val = ws.cell(row=r, column=2).value
        if val and 'Nama' in str(val):
            nama_row = r
            break
            
    if nama_row:
        signature_start = nama_row - 1
        target_signature_start = last_data_row + 3
        
        selisih = target_signature_start - signature_start
        if selisih != 0:
            ws.move_range(f"A{signature_start}:H{signature_start + 15}", rows=selisih, cols=0)

        for r in range(last_data_row + 1, target_signature_start):
            for c in range(1, 8):
                ws.cell(row=r, column=c).value = None
                ws.cell(row=r, column=c).border = Border()
                ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(fill_type=None)
                
        # Update Tanggal Tanda Tangan & Nama
        for r in range(target_signature_start, target_signature_start + 15):
            val = ws.cell(row=r, column=2).value
            if val and 'Tanggal' in str(val):
                ws.cell(row=r, column=3).value = ttd_date
                ws.cell(row=r, column=6).value = ttd_date
            if val and 'Nama' in str(val):
                ws.cell(row=r, column=3).value = nama_pegawai

    wb.save(file_output)
    return True, f"Sukses! Laporan disimpan di: {os.path.basename(file_output)}"

def main():
    print("=========================================")
    print("  GENERATOR LAPORAN ABSENSI (Terminal)   ")
    print("=========================================\n")
    
    default_csv = 'ekspor_csv Mei 26.csv'
    csv_input = input(f"1. Masukkan file CSV [Tekan Enter untuk '{default_csv}']: ").strip()
    csv_file = csv_input if csv_input else default_csv
    
    if not os.path.exists(csv_file):
        print(f"ERROR: File '{csv_file}' tidak ditemukan!")
        return

    print("\nMembaca daftar pegawai dari CSV...")
    try:
        df = pd.read_csv(csv_file, sep=';')
    except Exception as e:
        print(f"ERROR: Gagal membaca CSV ({e})")
        return
        
    if 'nama' not in df.columns:
        print("ERROR: Kolom 'nama' tidak ditemukan di file CSV.")
        return
        
    names = df['nama'].dropna().unique().tolist()
    if not names:
        print("ERROR: Tidak ada data pegawai di CSV.")
        return
        
    print("2. Pilih Nama Pegawai:")
    for i, name in enumerate(names, 1):
        print(f"   [{i}] {name}")
        
    nama_pegawai = ""
    while True:
        try:
            choice = input(f"Masukkan nomor (1-{len(names)}): ").strip()
            idx = int(choice) - 1
            if 0 <= idx < len(names):
                nama_pegawai = names[idx]
                print(f" -> Terpilih: {nama_pegawai}")
                break
            else:
                print(" -> Nomor di luar rentang, coba lagi.")
        except ValueError:
            print(" -> Harap masukkan angka yang valid.")
            
    print("\n3. Lengkapi Data (Tekan Enter untuk menggunakan default)")
    default_opd = "Dinas Komunikasi dan Informatika"
    opd_in = input(f"   OPD [{default_opd}]: ").strip()
    opd = opd_in if opd_in else default_opd
    
    default_projek = "Open SID (Sistem Informasi Desa) Kab. Badung"
    projek_in = input(f"   Nama Projek [{default_projek}]: ").strip()
    projek = projek_in if projek_in else default_projek
    
    default_role = "Full Stack Web Developer"
    role_in = input(f"   Role [{default_role}]: ").strip()
    role = role_in if role_in else default_role

    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Laporan Absensi Maret 2026.xlsx')
    if not os.path.exists(template_path):
        print(f"\nERROR: Template '{template_path}' tidak ditemukan!")
        return

    # Extract short name for filename
    words = nama_pegawai.split(' ')
    if len(words) > 1 and len(words[0]) <= 2:
        first_name = words[1]
    else:
        first_name = words[0]
        
    clean_name = re.sub(r'[^a-zA-Z0-9]', '', first_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_name = f"Laporan_Absensi_{clean_name}_{timestamp}.xlsx"
    file_output = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_name)

    print("\nMemproses laporan... Mohon tunggu.")
    try:
        success, msg = generate_laporan(
            df=df,
            template_excel=template_path,
            file_output=file_output,
            nama_pegawai=nama_pegawai,
            opd=opd,
            projek=projek,
            role=role
        )
        if success:
            print(f"\n{msg}\n")
        else:
            print(f"\nERROR: {msg}\n")
    except Exception as e:
        print(f"\nFATAL ERROR: {str(e)}\n")

if __name__ == "__main__":
    main()
