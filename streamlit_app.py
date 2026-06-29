import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
import shutil
import zipfile
import re
import os
from openpyxl.styles import Alignment, Border, Side, Font

# Konfigurasi Halaman Dasar
st.set_page_config(page_title="Generator Absensi", page_icon="📅", layout="centered")

# ==========================================
# 1. CORE LOGIC
# ==========================================

def strip_data_validations(excel_path):
    temp_path = excel_path + ".tmp"
    with zipfile.ZipFile(excel_path, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w') as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/'):
                    content_str = content.decode('utf-8')
                    content_str = re.sub(r'<dataValidations.*?</dataValidations>', '', content_str, flags=re.DOTALL)
                    content = content_str.encode('utf-8')
                zout.writestr(item, content)
    shutil.move(temp_path, excel_path)

def generate_excel_file(df, template_excel, file_output, nama_pegawai, opd, projek, role, format_durasi="Lengkap"):
    data_saya = df[df['nama'].str.lower() == nama_pegawai.lower()]

    if data_saya.empty:
        return False, f"Maaf, nama '{nama_pegawai}' tidak ditemukan di file CSV."

    hasil_rekap = []
    kolom_tanggal = df.columns[2:]

    for tanggal_str in kolom_tanggal:
        absen = data_saya.iloc[0][tanggal_str]
        tanggal_str_bersih = str(tanggal_str).strip()
        
        try:
            tanggal_obj = datetime.strptime(tanggal_str_bersih, '%d-%m-%Y')
            nama_hari = tanggal_obj.strftime('%A')
            tanggal_format_baru = tanggal_obj.strftime('%Y-%m-%d')
        except ValueError:
            continue

        jam_masuk = '-'
        jam_pulang = '-'
        durasi_kerja = '-'
        keterangan = '-'

        if nama_hari == 'Saturday':
            keterangan = 'Sabtu'
        elif nama_hari == 'Sunday':
            keterangan = 'Minggu'
        else:
            if pd.notna(absen) and str(absen).strip() != '':
                try:
                    if ' - ' in str(absen):
                        jam_masuk, jam_pulang = str(absen).split(' - ')
                        waktu_masuk = datetime.strptime(jam_masuk, '%H:%M:%S')
                        waktu_pulang = datetime.strptime(jam_pulang, '%H:%M:%S')
                        selisih = waktu_pulang - waktu_masuk
                        jam = selisih.seconds // 3600
                        menit = (selisih.seconds // 60) % 60
                        
                        if format_durasi == "Ringkas":
                            jam_ringkas = max(8, jam)
                            durasi_kerja = str(jam_ringkas)
                        else:
                            durasi_kerja = f'{jam} jam {menit} menit'
                            
                        keterangan = 'Hadir'
                    else:
                        jam_masuk = str(absen).strip()
                        keterangan = 'Format Error / Pulang Kosong'
                except Exception as e:
                    keterangan = 'Format Error/Tidak Lengkap'
            else:
                keterangan = 'Tidak Hadir / Cuti'

        hasil_rekap.append({
            'Tanggal(s)': tanggal_format_baru,
            'Jam Masuk': jam_masuk,
            'Jam Pulang': jam_pulang,
            'Durasi Kerja': durasi_kerja,
            'Keterangan': keterangan
        })

    ttd_date = datetime.now()

    # Pastikan file template asli ada
    if not os.path.exists(template_excel):
        return False, f"Error: Template {template_excel} tidak ditemukan di server!"

    shutil.copy(template_excel, file_output)
    strip_data_validations(file_output)

    wb = openpyxl.load_workbook(file_output)
    ws = wb.active

    # Update Biodata
    ws.cell(row=7, column=3, value=opd)
    ws.cell(row=7, column=6, value=nama_pegawai)
    ws.cell(row=8, column=3, value=projek)
    ws.cell(row=8, column=6, value=role)

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    regular_font = Font(name='Arial', size=11, bold=False)
    red_font = Font(name='Arial', size=11, bold=False, color="FF0000")
    left_align = Alignment(horizontal='left')

    for r in range(11, 45):
        for c in range(2, 7):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).border = Border()
            ws.cell(row=r, column=c).font = regular_font
            ws.cell(row=r, column=c).alignment = left_align
            
    baris_sekarang = 11
    for data_hari in hasil_rekap:
        ws.cell(row=baris_sekarang, column=2, value=data_hari['Tanggal(s)']).border = thin_border
        ws.cell(row=baris_sekarang, column=3, value=data_hari['Jam Masuk']).border = thin_border
        ws.cell(row=baris_sekarang, column=4, value=data_hari['Jam Pulang']).border = thin_border
        ws.cell(row=baris_sekarang, column=5, value=data_hari['Durasi Kerja']).border = thin_border
        ws.cell(row=baris_sekarang, column=6, value=data_hari['Keterangan']).border = thin_border
        
        keterangan = data_hari['Keterangan']
        is_red = (keterangan in ['Sabtu', 'Minggu', 'Tidak Hadir / Cuti'])
        current_font = red_font if is_red else regular_font

        for c in range(2, 7):
            ws.cell(row=baris_sekarang, column=c).font = current_font
            ws.cell(row=baris_sekarang, column=c).alignment = left_align
            
        baris_sekarang += 1

    last_data_row = baris_sekarang - 1
    
    nama_row = None
    for r in range(last_data_row, last_data_row + 20):
        val = ws.cell(row=r, column=2).value
        if val and 'Nama' in str(val):
            nama_row = r
            break
            
    if nama_row:
        signature_start = nama_row - 1
        target_signature_start = last_data_row + 3
        
        selisih = target_signature_start - signature_start
        if selisih != 0:
            ws.move_range(f"A{signature_start}:H{signature_start + 15}", rows=selisih, cols=0)

        for r in range(last_data_row + 1, target_signature_start):
            for c in range(1, 8):
                ws.cell(row=r, column=c).value = None
                ws.cell(row=r, column=c).border = Border()
                ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(fill_type=None)
                
        # Update Tanggal Tanda Tangan & Nama
        for r in range(target_signature_start, target_signature_start + 15):
            val = ws.cell(row=r, column=2).value
            if val and 'Tanggal' in str(val):
                ws.cell(row=r, column=3).value = ttd_date
                ws.cell(row=r, column=6).value = ttd_date
            if val and 'Nama' in str(val):
                ws.cell(row=r, column=3).value = nama_pegawai

    wb.save(file_output)
    return True, "Berhasil!"

# ==========================================
# 2. UI STREAMLIT
# ==========================================

st.title("📑 Generator Laporan Absensi")
st.markdown("Ubah data CSV mentah menjadi laporan Excel resmi dengan satu klik.")

# File template patokan harus ada di folder repo saat dideploy
template_path = "Laporan Absensi Maret 2026.xlsx"

# Widget Unggah File
uploaded_file = st.file_uploader("1. Unggah File CSV Absensi", type=["csv"])

if uploaded_file is not None:
    try:
        # Pandas secara otomatis bisa membaca dari Streamlit memory buffer
        df = pd.read_csv(uploaded_file, sep=';')
        
        if 'nama' in df.columns:
            names = df['nama'].dropna().unique().tolist()
            st.success("File CSV valid dan berhasil dibaca!")
            
            # Buat Layout Grid agar cantik
            col1, col2 = st.columns([1, 1])
            
            with col1:
                nama_pegawai = st.selectbox("2. Pilih Nama Pegawai", names)
                opd = st.text_input("3. OPD", value="Dinas Komunikasi dan Informatika")
                
            with col2:
                projek = st.text_input("4. Nama Projek", value="Open SID (Sistem Informasi Desa) Kab. Badung")
                role = st.text_input("5. Role", value="Full Stack Web Developer")
            
            # Format Durasi
            st.write("6. Pengaturan Laporan")
            format_durasi = st.radio("Format Durasi Kerja", ["Lengkap (Misal: 9 jam 15 menit)", "Ringkas (Hanya Angka Jam, Minimal 8)"])
            format_internal = "Ringkas" if "Ringkas" in format_durasi else "Lengkap"
            
            # Tombol Eksekusi
            if st.button("🚀 Generate Laporan Excel", type="primary", use_container_width=True):
                with st.spinner("Sedang memproses laporan..."):
                    # Buat nama file temp output di server Streamlit
                    temp_out = "temp_output.xlsx"
                    
                    try:
                        success, msg = generate_excel_file(
                            df=df,
                            template_excel=template_path,
                            file_output=temp_out,
                            nama_pegawai=nama_pegawai,
                            opd=opd,
                            projek=projek,
                            role=role,
                            format_durasi=format_internal
                        )
                        
                        if success:
                            # Baca ulang file Excel dari server ke memori (Byte) untuk dikirim ke User Download
                            with open(temp_out, "rb") as f:
                                excel_data = f.read()
                                
                            # Siapkan nama file cantik untuk didownload
                            words = nama_pegawai.split(' ')
                            first_name = words[1] if (len(words) > 1 and len(words[0]) <= 2) else words[0]
                            clean_name = re.sub(r'[^a-zA-Z0-9]', '', first_name)
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                            download_name = f"Laporan_Absensi_{clean_name}_{timestamp}.xlsx"
                            
                            st.success("Yeay! Laporan berhasil dibuat! Silakan klik tombol di bawah untuk mengunduh.")
                            
                            # Tombol untuk mendownload (menyalurkan Bytes dari RAM ke Browser pengguna)
                            st.download_button(
                                label="⬇️ Download Laporan Anda",
                                data=excel_data,
                                file_name=download_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        else:
                            st.error(msg)
                    except Exception as e:
                        st.error(f"Terjadi kesalahan saat generate: {e}")
                    finally:
                        # Membersihkan sampah temporary di server setelah file disajikan
                        if os.path.exists(temp_out):
                            os.remove(temp_out)
                            
        else:
            st.error("Gagal! File CSV tidak memiliki kolom 'nama'. Pastikan file dari mesin absensi tidak diubah.")
            
    except Exception as e:
        st.error(f"Gagal membaca CSV: {e}")
